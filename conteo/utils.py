import openpyxl
from openpyxl.utils import get_column_letter
from django.core.mail import EmailMessage
from .models import ConteoDiario
import tempfile
import os

def generar_y_enviar_excel(sucursal, usuario, email_destino):
    # Crear un nuevo libro de trabajo
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"Conteo de {sucursal.nombre}"  # Usar nombre de la sucursal

    # Agregar la sucursal y el usuario en las primeras filas
    sheet['A1'] = 'Sucursal:'
    sheet['B1'] = sucursal.nombre  # Mostrar el nombre
    sheet['A2'] = 'Usuario:'
    sheet['B2'] = usuario.get_full_name()

    # Crear los encabezados
    headers = ['Producto', 'Cantidad Contada', 'Fecha de Conteo']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}4'] = header

    # Filtrar los conteos usando la instancia de la sucursal
    conteos = ConteoDiario.objects.filter(sucursal=sucursal, usuario=usuario)

    # Llenar el archivo Excel con los datos
    for row_num, conteo in enumerate(conteos, start=5):
        sheet[f'A{row_num}'] = conteo.producto.nombre
        sheet[f'B{row_num}'] = conteo.cantidad_contada
        sheet[f'C{row_num}'] = conteo.fecha_conteo.strftime('%Y-%m-%d')

    # Usar un archivo temporal para guardar el Excel
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        workbook.save(tmp.name)
        tmp_path = tmp.name

    # Enviar el archivo Excel por correo electrónico
    email = EmailMessage(
        subject=f'Reporte de Conteo Diario - {sucursal.nombre}',
        body=f'Adjunto encontrarás el reporte del conteo realizado por {usuario.get_full_name()}.',
        from_email='luchoviteri1990@gmail.com',
        to=[email_destino],
    )
    email.attach_file(tmp_path)
    email.send()

    # Eliminar los registros de conteo después de enviar el correo
    conteos.delete()

    # Eliminar el archivo temporal
    os.remove(tmp_path)
